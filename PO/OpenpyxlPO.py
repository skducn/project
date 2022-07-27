# -*- coding: utf-8 -*-
# *********************************************************************
# Author        : John
# Date          : 2020-12-8
# Description   : openpyxl 对象层
# openpyxl 官网：http://openpyxl.readthedocs.org/en/latest/
# 支持.xlsx / .xlsm / .xltx / .xltm格式的文件
# 首行、首列 是 （1,1）而不是（0,0）
# NULL空值对应于python中的None，表示这个cell里面没有数据。
# numberic数字型，统一按照浮点数来进行处理，对应于python中的float
# string字符串型，对应于python中的unicode
# openpyxl 会将整个xlsx读入到内存中，方便处理。
# openpyxl 操作大文件时可使用 Optimized reader 和 Optimized writer 两种模式，它们提供了流式的接口，速度更快，使我们可以用常量级的内存消耗来读取和写入无限量的数据。
# Optimized reader，打开文件使用use_iterators=True参数，如：wb = load_workbook(filename = 'haggle.xlsx',use_iterators=True)
# openpyxl 读取大数据的效率没有 xlrd 高
# openpyxl 与 xlsxwriter xlrd xlwt xlutils 的比较，这些库都不支持 excel 写操作，一般只能将原excel中的内容读出、做完处理后，再写入一个新的excel文件。

# todo: 使用方法
# 参考：openpyxl常用模块用法 https://www.debug8.com/python/t_41519.html
# 基础使用方法：https://blog.csdn.net/four91/article/details/106141274
# 高级使用方法：https://blog.csdn.net/m0_47590417/article/details/119082064

# todo: 安装包
# pip3 install openpyxl == 3.0.0
# 注意！：其他版本（如3.0.2使用中会报错）如有报错，请安装3.0.0

# todo: 报错
# 如：File "src\lxml\serializer.pxi", line 1652, in lxml.etree._IncrementalFileWriter.write TypeError: got invalid input value of type <class 'xml.etree.ElementTree.Element'>, expected string or Element
# 解决方法: pip uninstall lxml   及更新 openpyxl 版本，3.0.7以上

# todo: 乱码
# gb2312 文字编码，在读取后会显示乱码，需转换成 Unicode

# todo: 颜色
# 颜色码对照表（RGB与十六进制颜色码互转） https://www.sioe.cn/yingyong/yanse-rgb-16/
# 绿色 = 00E400，黄色 = FFFF00，橙色 = FF7E00，红色 = FF0000，粉色 = 99004C，褐色 =7E0023,'c6efce = 淡绿', '006100 = 深绿'，'ffffff=白色', '000000=黑色'，'ffeb9c'= 橙色

# todo: 表格列 A，B，C 与 1，2，3 互转
from openpyxl.utils import get_column_letter,column_index_from_string
# get_column_letter(2)  # 'B'
# column_index_from_string('B')  # 2
# *********************************************************************

from openpyxl import load_workbook
import openpyxl, sys, platform, os
import openpyxl.styles
from openpyxl.styles import Font, PatternFill, GradientFill, Border, Side, Alignment, Protection, Alignment
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import date
from time import sleep
import psutil
from PO.ColorPO import *
Color_PO = ColorPO()
from PO.CharPO import *
Char_PO = CharPO()
from PO.SysPO import *
Sys_PO = SysPO()
from PO.MysqlPO import *

'''
1.1 新建  newExcel("./OpenpyxlPO/newfile2.xlsx", "mySheet1", "mySheet2", "mySheet3") 
1.2 打开 open()
1.3 获取所有工作表 getSheets()
1.4 操作工作表 sh()
1.5 切换工作表 switchSheet("Sheet2") 
1.6 添加工作表(不覆盖)  addSheet("Sheet1")
1.7 添加工作表(覆盖) addSheetCover("Sheet1", 1) 
1.8 删除工作表  delSheet("Sheet1")
1.9 保存 save()

2.0 
插入一行 insertRows(3) 
插入一列 insertCols(3)

2.1 设置单元格值 setCellValue(1, 6, "jinhao")
2.2 设置整行值  setRowValue({7: ["你好", 12345, "7777"], 8: ["44", None, "777777777"]})
2.3 设置整列值  setColValue({"A": ["k1", 666, "777"], "F": ["4456", None, "888"]}, -1)
2.4 追加整行值 addOnRowValue([['姓名', '电话', '成绩', '学科'], ['毛泽东', 15266606298, 14, '化学'], ['周恩来', 15201077791, 78, '美术']])


2.5 设置单元格行高与列宽 setCellDimensions(3, 30, 'f', 30) //设置第三行行高30，第f列列宽50
2.6 设置工作表所有单元格的行高与列宽 setAllCellDimensions(30, 20) //设置所有单元格高30，宽50
2.7 设置所有单元格自动换行 setAllWordWrap()
2.8 设置冻结首行 setFreeze('A2'）
2.9 设置单元格对齐样式  setCellAlignment(5, 4, 'center', 'center')
2.9.2 设置单行多列对齐样式 setRowColAlignment(5, [1,4], 'center', 'center')
2.9.3 设置所有单元格对齐样式 setAllCellAlignment('center', 'center')
2.10 设置筛选列  setFilterCol("all") # 全部筛选, setFilterCol("") # 取消筛选 , setFilterCol("A2") # 对A2筛选 
2.11 设置单元格字体（字体、字号、粗斜体、下划线、颜色） setCellFont(1, 1, name=u'微软雅黑', size=16, bold=True, italic=True, color="000000")
2.11.2 设置单行多列字体  setRowColFont(1, [1, 5])
2.11.3 设置所有单元格字体  setAllCellFont(color="000000")
2.12 设置单元格边框 setBorder(1, 2, left = ['thin','ff0000'], right = ['thick','ff0000'], top = ['thin','ff0000'],bottom = ['thick','ff0000'])
2.13 设置单元格填充背景色 setPatternFill(2, 2, 'solid', '006100')
2.14 设置单元格填充渐变色 setGradientFill(3, 3, stop=["FFFFFF", "99ccff", "000000"])
2.15 设置单元格背景色 setCellColor(1, 1, "ff0000")  # 第1行第一列设置红色
2.15.2 设置单行多列背景色 setRowColColor(5, ['b', 'd'], "ff0000")
2.15.3 设置所有单元格背景色 setAllCellColor("ff0000")
2.16 设置整行(可间隔)背景色  setRowColor(3, 1, "ff0000")  # 从第3行开始每隔1行颜色标红色
2.17 设置整列(可间隔)背景色  setColColor(2, 1, "ff0000")  # 从第2列开始每隔1列设置颜色为红色
2.18 设置工作表背景颜色 setSheetColor("FF0000")

3.1 获取总行数和总列数 getRowCol()
3.2 获取单元格的值 getCellValue()
3.3 获取单行数据 getOneRowValue(2)
3.4 获取每行数据 getRowValue()
3.5 获取每列数据 getColValue()
3.6 获取指定列的行数据 getRowValueByCol([1, 2, 4], -1))   # 获取最后一个工作表的第1，2，4列的行数据
3.7 获取某些列的列数据(可忽略多行) getColValueByCol([1, 3], [1, 2]))   # 获取第二列和第四列的列值，并忽略第1，2行的行值。
3.8 获取单元格的坐标 getCoordinate(2, 5))   # E2
3.9 获取所有数据的坐标 getDimensions())  # A1:E17


4.1 清空行 clsRow(2)  # 清空第2行
4.2 清空列 clsCol(2)  # 清空第2列
4.2.1 清空列保留标题 clsColRetainTitle(2)  # 清空第2列
4.3 删除连续行 delSeriesRow(2, 3)  # 删除从第二行开始连续三行数据 （即删除2，3，4行）
4.4 删除连续列 delSeriesCol(2, 3)  # 删除从第二列开始连续三列数据 （即删除2，3，4列）

5.1 两表比较，获取差异内容（两表标题与行数必须一致）getDiffValueByCmp(Openpyxl_PO.getRowValue("Sheet2"), Openpyxl_PO2.getRowValue("Sheet2"))
5.2 两工作表比较，对差异内容标注颜色 setColorByDiff("Sheet1", "Sheet2")
 
6 移动范围数据 moveValue(rows, cols, 'C1:D2')


'''

class OpenpyxlPO():

    def __init__(self, file):

        self.file = file
        self.wb = openpyxl.load_workbook(self.file)
        # self.wb.sheetnames
        # self.wb.active  # 获取当前活跃的Worksheet对象
        # print(self.wb.active)  # <Worksheet "北京">
        # self.wb.worksheets  # 以列表的形式返回所有的Worksheet对象，如：[<Worksheet "北京">, <Worksheet "mySheet1">, <Worksheet "上海">]
        # sh2 = self.wb['上海']
        # print(sh2)
        # self.wb.encoding  # 获取文档的字符集编码
        # self.wb.properties  # 获取文档的元数据，如标题，创建者，创建日期等
        # self.wb.active = 0  # 通过索引值设置当前活跃的worksheet


    # todo [工作表]

    def newExcel(self, varFileName, *varSheetName):

        # 1.1 新建excel(覆盖)
        # Openpyxl_PO.newExcel("d:\\444.xlsx")  # 新建excel默认一个Sheet1工作表
        # Openpyxl_PO.newExcel("d:\\444.xlsx", "mySheet1", "mySheet2","mySheet3")  # 新建excel生成三个工作表，默认在第一个mySheet1表。
        # 注意：如果文件已存在则会先删除后再新建。
        wb = openpyxl.Workbook()
        ws = wb.active
        if len(varSheetName) == 0:
            ws.title = "Sheet1"
        else:
            ws.title = varSheetName[0]
        for i in range(1, len(varSheetName)):
            wb.create_sheet(varSheetName[i])
        wb.save(varFileName)



    def open(self, otherFile=0):

        # 1.2 打开
        if platform.system() == 'Darwin':
            if otherFile != 0:
                os.system("open " + otherFile)
            else:
                os.system("open " + self.file)
        if platform.system() == 'Windows':
            if otherFile != 0 :
                os.system("start " + otherFile)
            else:
                os.system("start " + self.file)


    def getSheets(self):

        # 1.3 获取所有工作表，如 ['mySheet1', 'mySheet2', 'mySheet3']
        return self.wb.sheetnames


    def sh(self, varSheet):

        # 1.4 操作工作表
        if isinstance(varSheet, int):
            sh = self.wb[self.wb.sheetnames[varSheet]]
            return sh
        elif isinstance(varSheet, str):
            sh = self.wb[varSheet]
            return sh
        else:
            exit(0)


    def switchSheet(self, varSheet):

        # 1.5 切换工作表
        # switchSheet("Sheet2")
        return self.wb[varSheet]


    def addSheet(self, varSheetName, varIndex=0):

        # 1.6 添加工作表(不覆盖)
        # Openpyxl_PO.addSheet("mysheet1")  # 默认在第一个位置上添加工作表
        # Openpyxl_PO.addSheet("mysheet1", 99)   # 当index足够大时，则在最后一个位置添加工作表
        # Openpyxl_PO.addSheet("mysheet1", -1)   # 倒数第二个位置添加工作表
        # 注意：如果工作表名已存在，则不添加工作表，即保留原工作表。
        sign = 0
        for i in self.wb.sheetnames:
            if i == varSheetName:
                sign = 1
                break
        if sign == 0:
            self.wb.create_sheet(title=varSheetName, index=varIndex)
            self.save()


    def addSheetCover(self, varSheetName, varIndex=0):

        # 1.7 添加工作表(覆盖)
        # Openpyxl_PO.addSheetCover("mySheet1")
        # Openpyxl_PO.addSheetCover("mySheet1", 0 )  # 在第一个工作表前添加工作表
        # Openpyxl_PO.addSheetCover("mySheet2",99)   # 在第99个位置添加工作表
        # Openpyxl_PO.addSheetCover("mySheet3", -1)   # 在倒数第二个位置添加工作表。
        for i in self.wb.sheetnames:
            if i == varSheetName:
                del self.wb[i]
                break
        self.wb.create_sheet(title=varSheetName, index=varIndex)
        self.save()


    def delSheet(self, varSheetName):

        # 1.8 删除工作表
        # Openpyxl_PO.delSheet("mySheet1")
        # 注意:如果工作表只有1个，则不能删除。
        if len(self.wb.sheetnames) > 1:
            for i in self.wb.sheetnames:
                if i == varSheetName:
                    del self.wb[i]
                    self.save()
        else:
            print("[warning], excel必须保留1个工作表！")


    def save(self):

        # 1.9 保存
        self.wb.save(self.file)



    # todo [设置]


    def insertRows(self, row, moreRow=0, varSheet=0):

        # 2.0.1 插入一行
        sh = self.sh(varSheet)
        sh.insert_rows(idx=row, amount=moreRow)
        self.save()


    def insertCols(self, col, moreCol=0, varSheet=0):

        # 2.0.2 插入一列
        sh = self.sh(varSheet)
        sh.insert_cols(idx=col, amount=moreCol)
        self.save()

    def setCellValue(self, varRow, varCol, varContent, varSheet=0):

        # 2.1 设置单元格值
        sh = self.sh(varSheet)
        sh.cell(row=varRow, column=varCol, value=varContent)
        self.save()


    def setRowValue(self, d_var, varSheet=0):

        # 2.2 设置整行值
        # Openpyxl_PO.setRowValue({7:[1,2,3],8:["44",66]})  # 对第7行第9行分别写入内容
        # Openpyxl_PO.setRowValue({7: ["你好", 12345, "7777"], 8: ["44", None, "777777777"]}, -1)  # 对最后一个sheet表，对第7，8行分别写入内容，如遇None则跳过该单元格
        sh = self.sh(varSheet)
        for k, v in d_var.items():
            for i in range(len(v)):
                if v[i] != None:
                    sh.cell(row=k, column=i + 1, value=str(v[i]))
        self.save()


    def setColValue(self, d_var, varSheet=0):

        # 2.3 设置整列值
        # Openpyxl_PO.setColValue({"A": ["k1", 666, "777"], "F": ["name", None, "888"]}, -1)
        sh = self.sh(varSheet)
        for k, v in d_var.items():
            for i in range(len(v)):
                if v[i] != None:
                    sh.cell(row=i + 1, column=column_index_from_string(k), value=v[i])
        self.save()


    def addOnRowValue(self, data, varSheet=0):

        # 2.4 追加整行值
        # addOnRowValue([['姓名', '电话', '成绩', '学科'], ['毛泽东', 15266606298, 14, '化学'], ['周恩来', 15201077791, 78, '美术']])
        sh = self.sh(varSheet)
        for r in range(len(data)):
            sh.append(data[r])
        self.save()


    def setCellDimensions(self, row, rowQty, col, colQty, varSheet=0):

        # 2.5 设置单元格行高与列宽
        # Openpyxl_PO.setCellDimensions(3, 30, 'f', 50) //第三行行高30，第f列列宽50
        sh = self.sh(varSheet)
        sh.row_dimensions[row].height = rowQty  # 行高
        sh.column_dimensions[col].width = colQty  # 列宽
        self.save()


    def setRowColDimensions(self, row, rowQty, l_col, colQty, varSheet=0):

        # 2.5.2 设置单行多列行高与列宽
        # Openpyxl_PO.setRowColDimensions(5, 30, ['f', 'h'], 30)  #
        sh = self.sh(varSheet)
        cols = sh.max_column
        sh.row_dimensions[row].height = rowQty  # 行高
        if l_col == "all":
            for i in range(1, cols + 1):
                sh.column_dimensions[get_column_letter(i)].width = colQty  # 列宽
        else:
            # print(column_index_from_string(l_col[0]))  # 6
            # print(column_index_from_string(l_col[1]))  # 8
            for i in range(column_index_from_string(l_col[0]), int(column_index_from_string(l_col[1])) + 1):
                sh.column_dimensions[get_column_letter(i)].width = colQty  # 列宽
        self.save()


    def setAllCellDimensions(self, rowQty, colQty, varSheet=0):

        # 2.6 设置所有单元格的行高与列宽
        sh = self.sh(varSheet)
        rows = sh.max_row
        columns = sh.max_column
        for i in range(1, rows+1):
            sh.row_dimensions[i].height = rowQty  # 行高
        for i in range(1, columns+1):
            sh.column_dimensions[get_column_letter(i)].width = colQty  # 列宽
        self.save()


    def setAllWordWrap(self, varSheet=0):

        # 2.7 设置所有单元格自动换行
        sh = self.sh(varSheet)
        # print(list(sh._cells.keys())) # [(1, 1), (1, 2), (1, 3), (2, 1), (2, 2), (2, 3), (3, 1), (3, 2), (3, 3), (4, 1), (4, 2), (4, 3), (5, 1), (5, 2), (5, 3), (6, 1), (6, 2), (6, 3), (7, 1), (7, 2), (7, 3)]
        for key in list(sh._cells.keys()):
            sh._cells[key].alignment = Alignment(wrapText=True)
        self.save()


    def setFreeze(self, coordinate, varSheet=0):

        # 2.8 冻结窗口
        # setFreeze('A2'）
        sh = self.sh(varSheet)
        sh.freeze_panes = coordinate
        self.save()


    def setCellAlignment(self, row, col, horizontal='center', vertical='center', text_rotation=0, wrap_text=False, varSheet=0):

        # 2.9 设置单元格对齐样式
        # Openpyxl_PO.setCellAlignment(5, 4, 'center', 'top')
        # Openpyxl_PO.setCellAlignment(5, "f", 'center', 'top')
        # Alignment(horizonta水平对齐模式, vertical=垂直对齐模式, text_rotation=旋转角度, wrap_text=是否自动换行)
        # horizontal = ("general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed",)
        # vertical = ("top", "center", "bottom", "justify", "distributed")
        sh = self.sh(varSheet)
        if isinstance(col, int):
            sh.cell(row=row, column=col).alignment = Alignment(horizontal=horizontal, vertical=vertical, text_rotation=text_rotation, wrap_text=wrap_text)
        else:
            sh.cell(row, column_index_from_string(col)).alignment = Alignment(horizontal=horizontal, vertical=vertical, text_rotation=text_rotation, wrap_text=wrap_text)
        self.save()


    def setRowColAlignment(self, row, l_col, horizontal='center', vertical='center', text_rotation=0, wrap_text=False, varSheet=0):

        # 2.9.2 设置单行多列对齐样式
        # Openpyxl_PO.setRowColAlignment(1, [4, 6], 'center', 'center')  # 第一行第四五六列居中
        # Openpyxl_PO.setRowColAlignment(9, "all", 'center', 'center')  # 第九行全部居中
        # Alignment(horizonta水平对齐模式, vertical=垂直对齐模式, text_rotation=旋转角度, wrap_text=是否自动换行)
        # horizontal = ("general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed",)
        # vertical = ("top", "center", "bottom", "justify", "distributed")
        sh = self.sh(varSheet)
        cols = sh.max_column
        if l_col == "all":
            for i in range((cols)):
                sh.cell(row=row, column=i+1).alignment = Alignment(horizontal=horizontal, vertical=vertical, text_rotation=text_rotation, wrap_text=wrap_text)
        else:
            for i in range(column_index_from_string(l_col[0]), int(column_index_from_string(l_col[1])) + 1):
                sh.cell(row=row, column=i).alignment = Alignment(horizontal=horizontal, vertical=vertical, text_rotation=text_rotation, wrap_text=wrap_text)
        self.save()


    def setAllCellAlignment(self, horizontal='center', vertical='center', text_rotation=0, wrap_text=False, varSheet=0):

        # 2.9.3 设置所有单元格对齐样式
        # Openpyxl_PO.setAllCellAlignment('center', 'center')
        # Alignment(horizonta水平对齐模式, vertical=垂直对齐模式, text_rotation=旋转角度, wrap_text=是否自动换行)
        # horizontal = ("general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed",)
        # vertical = ("top", "center", "bottom", "justify", "distributed")
        sh = self.sh(varSheet)
        rows = sh.max_row
        cols = sh.max_column
        for r in range(rows):
            for c in range(cols):
                sh.cell(row=r+1, column=c+1).alignment = Alignment(horizontal=horizontal, vertical=vertical, text_rotation=text_rotation, wrap_text=wrap_text)
        self.save()


    def setFilterCol(self, varCell="all", varSheet=0):

        # 2.10 设置筛选列
        # setFilterCol("all")  # 全部筛选
        # setFilterCol("")  # 取消筛选
        # setFilterCol("A2") # 对A2筛选
        sh = self.sh(varSheet)
        if varCell == "all":
            sh.auto_filter.ref = sh.dimensions
        elif varCell == "":
            sh.auto_filter.ref = None
        else:
            sh.auto_filter.ref = varCell
        self.save()


    def setCellFont(self, row, col, name=u'微软雅黑', size=16, bold=False, italic=False, color=None, varSheet=0):

        # 2.11 设置单元格字体（字体、字号、粗斜体、下划线、颜色）
        # setCellFont(1, 1)
        # setCellFont(1, "f")
        # setCellFont(1, "f", size=16, bold=True, color="ff0000")
        sh = self.sh(varSheet)
        if isinstance(col, int):
            sh.cell(row, col).font = Font(name=name, size=size, bold=bold, italic=italic, color=color)
        else:
            sh.cell(row, column_index_from_string(col)).font = Font(name=name, size=size, bold=bold, italic=italic, color=color)
        self.save()


    def setRowColFont(self, row, l_col, name=u'微软雅黑', size=16, bold=False, italic=False, color="000000", varSheet=0):

        # 2.11.2 设置单行多列字体（字体、字号、粗斜体、下划线、颜色）
        # setRowColFont(1, ["e", "h"])
        # setRowColFont(1, "all", color="000000")
        sh = self.sh(varSheet)
        cols = sh.max_column
        if l_col == "all":
            for i in range((cols)):
                sh.cell(row=row, column=i + 1).font = Font(name=name, size=size, bold=bold, italic=italic, color=color)
        else:
            for i in range(column_index_from_string(l_col[0]), int(column_index_from_string(l_col[1])) + 1):
                sh.cell(row=row, column=i).font = Font(name=name, size=size, bold=bold, italic=italic, color=color)
        self.save()


    def setAllCellFont(self, name=u'微软雅黑', size=16, bold=False, italic=False, color="000000", varSheet=0):

        # 2.11.2 设置所有单元格字体（字体、字号、粗斜体、下划线、颜色）
        # setAllCellFont()
        sh = self.sh(varSheet)
        rows = sh.max_row
        cols = sh.max_column

        for r in range(rows):
            for c in range(cols):
                sh.cell(row=r + 1, column=c + 1).font = Font(name=name, size=size, bold=bold, italic=italic, color=color)
        self.save()


    def setBorder(self, row, col, left = ['thin','ff0000'], right = ['thick','ff0000'], top = ['thin','ff0000'],bottom = ['thick','ff0000'],  varSheet=0):

        '''
        2.12 设置单元格边框
        # 设置边框样式，上下左右边框
        Side(style=边线样式，color=边线颜色)
         * style 参数的种类： 'double, 'mediumDashDotDot', 'slantDashDot','dashDotDot','dotted','hair', 'mediumDashed, 'dashed', 'dashDot', 'thin','mediumDashDot','medium', 'thick'

        setBorder(1, 2, left = ['thin','ff0000'], right = ['thick','ff0000'], top = ['thin','ff0000'],bottom = ['thick','ff0000'])
        '''

        sh = self.sh(varSheet)
        border = Border(left=Side(style=left[0], color=left[1]), right=Side(style=right[0], color=right[1]),
                        top=Side(style=top[0], color=top[1]), bottom=Side(style=bottom[0], color=bottom[1]))
        sh.cell(row=row, column=col).border = border
        self.save()


    def setPatternFill(self, row, col, fill_type="solid", fgColor="99ccff", varSheet=0):

        '''
        2.13 设置单元格填充背景色
        patternType = {'lightVertical', 'mediumGray', 'lightGrid', 'darkGrid', 'gray125', 'lightHorizontal', 'gray0625','lightTrellis', 'darkUp', 'lightGray', 'darkVertical', 'darkGray', 'solid', 'darkTrellis', 'lightUp','darkHorizontal', 'darkDown', 'lightDown'}
        PatternFill(fill_type=填充样式，fgColor=填充颜色）
        setPatternFill(2, 2, 'solid', '006100')
        '''

        sh = self.sh(varSheet)
        pattern_fill = PatternFill(fill_type=fill_type, fgColor=fgColor)
        sh.cell(row=row, column=col).fill = pattern_fill
        self.save()
        # return PatternFill(patternType='' + patternType + '', fgColor='' + fgColor + '')  # 背景色


    def setGradientFill(self, row, col, stop=["FFFFFF", "99ccff", "000000"], varSheet=0):

        # 2.14 设置单元格填充渐变色
        # GradientFill(stop=(渐变颜色 1，渐变颜色 2……))
        # setGradientFill(3, 3, stop=["FFFFFF", "99ccff", "000000"])
        sh = self.sh(varSheet)
        gradient_fill = GradientFill(stop=(stop[0], stop[1], stop[2]))
        sh.cell(row=row, column=col).fill = gradient_fill
        self.save()


    def setCellColor(self, row, col, varColor=None, varSheet=0):

        # 2.15 设置单元格背景色
        # Openpyxl_PO.setCellColor(5, 1)  # 清除第5行第1列的背景色
        # Openpyxl_PO.setCellColor(5, "d")  # 清除第5行d列的背景色
        # Openpyxl_PO.setCellColor(5, 1, "ff0000")  # 设置第五行第1列设置红色
        # Openpyxl_PO.setCellColor(5, "e", "ff0000")  # 设置第五行e列设置红色
        # Openpyxl_PO.setCellColor(None, None)  # 清除所有背景色
        sh = self.sh(varSheet)
        rows = sh.max_row
        cols = sh.max_column
        # 清除表格里所有单元格的背景色
        if row == None and col == None:
            style = PatternFill(fill_type=None)
            for i in range(1, rows + 1):
                for j in range(1, cols + 1):
                    sh.cell(i, j).fill = style
        else:
            if varColor == None:
                style = PatternFill(fill_type=None)  # 消除单元格颜色
            else:
                style = PatternFill("solid", fgColor=varColor)
            if isinstance(col, int):
                sh.cell(row, col).fill = style
            else:
                sh.cell(row, column_index_from_string(col)).fill = style
        self.save()


    def setRowColColor(self, row, l_col, varColor, varSheet=0):

        # 2.15.2 设置单行多列背景色
        # Openpyxl_PO.setRowColColor(5, ['b', 'd'], "ff0000")
        # Openpyxl_PO.setRowColColor(5, "all", "ff0000")
        sh = self.sh(varSheet)
        cols = sh.max_column
        style = PatternFill("solid", fgColor=varColor)

        if l_col == "all":
            for i in range((cols)):
                sh.cell(row=row, column=i+1).fill = style
        else:
            for i in range(column_index_from_string(l_col[0]), int(column_index_from_string(l_col[1])) + 1):
                sh.cell(row=row, column=i).fill = style
        self.save()


    def setAllCellColor(self, varColor=None, varSheet=0):

        # 2.15.3 设置所有单元格背景色
        # setAllCellColor("ff0000")
        # setAllCellColor()
        sh = self.sh(varSheet)
        rows = sh.max_row
        cols = sh.max_column
        if varColor == None:
            style = PatternFill(fill_type=None)  # 消除单元格颜色
        else:
            style = PatternFill("solid", fgColor=varColor)

        for r in range(rows):
            for c in range(cols):
                sh.cell(row=r+1, column=c+1).fill = style
        self.save()


    def setRowColor(self, row, varSkip, varColor, varSheet=0):

        # 2.16 设置整行(可间隔)背景色
        # Openpyxl_PO.setRowColor(6, 1, "FF0000")
        sh = self.sh(varSheet)
        rows = sh.max_row
        cols = sh.max_column
        # 清除表格里所有单元格的背景色
        style = PatternFill(fill_type=None)
        for i in range(1, rows+1):
            for j in range(1, cols+1):
                sh.cell(i, j).fill = style
        style = PatternFill("solid", fgColor=varColor)
        for i in range(row, rows+1, varSkip+1):
            for j in range(1, cols+1):
                sh.cell(i, j).fill = style
        self.save()


    def setColColor(self, col, varSkip, varColor, varSheet=0):

        # 2.17 设置整列(可间隔)背景色
        # Openpyxl_PO.setColColor(6, 1, "FF0000")
        sh = self.sh(varSheet)
        rows = sh.max_row
        cols = sh.max_column
        style = PatternFill(fill_type=None)  # 消除单元格颜色
        for i in range(1, rows+1):
            for j in range(1, cols+1):
                sh.cell(i, j).fill = style
        style = PatternFill("solid", fgColor=varColor)
        for i in range(1, rows+1):
            for j in range(col, cols+1, varSkip+1):
                sh.cell(i, j).fill = style
        self.save()


    def setSheetColor(self, varColor, varSheet=0):

        # 2.18 设置工作表背景颜色
        # Openpyxl_PO.setSheetColor("FF0000")
        sh = self.sh(varSheet)
        sh.sheet_properties.tabColor = varColor
        self.save()



    # todo [获取]

    def getRowCol(self, varSheet=0):

        # 3.1 获取总行数和总列数
        # Openpyxl_PO.getRowCol()  # [4,3] //返回第1个工作表的总行数和总列数
        sh = self.sh(varSheet)
        rows = sh.max_row
        cols = sh.max_column
        return [rows, cols]


    def getCellValue(self, varRow, varCol, varSheet=0):

        # 3.2 获取单元格的值
        sh = self.sh(varSheet)
        cell_value = sh.cell(row=varRow, column=varCol).value
        return cell_value


    def getOneRowValue(self, varRow, varSheet=0):

        # 3.3 获取单行数据
        # getOneRowValue(2)
        list1 = []
        sh = self.sh(varSheet)
        row = [val for val in sh.rows][varRow]  # 获取第一行
        for cel in row:
            list1.append(cel.value)
        return list1


    def getRowValue(self, varSheet=0):

        # 3.4 获取每行数据
        # print(Openpyxl_PO.getRowValue())
        l_rowData = []  # 每行数据
        l_allData = []  # 所有行数据
        sh = self.sh(varSheet)
        for cases in list(sh.rows):
            for i in range(len(cases)):
                l_rowData.append(cases[i].value)
            l_allData.append(l_rowData)
            l_rowData = []
        return (l_allData)


    def getColValue(self, varSheet=0):

        # 3.5 获取每列数据
        # print(Openpyxl_PO.getColValue())
        l_colData = []  # 每列数据
        l_allData = []  # 所有行数据
        sh = self.sh(varSheet)
        for cases in list(sh.columns):
            for i in range(len(cases)):
                l_colData.append(cases[i].value)
            l_allData.append(l_colData)
            l_colData = []
        return (l_allData)


    def getRowValueByCol(self, l_varCol, varSheet=0):

        # 3.6 获取指定列的行数据
        # print(Openpyxl_PO.getRowValueByCol([1, 2, 4]))  # 获取第1，2，4列的行数据
        l_rowData = []  # 每行的数据
        l_allData = []  # 所有的数据
        sh = self.sh(varSheet)
        for row in range(1, sh.max_row + 1):
            for column in l_varCol:
                l_rowData.append(sh.cell(row, column).value)
            l_allData.append(l_rowData)
            l_rowData = []
        return l_allData


    def getColValueByCol(self, l_varCol, l_varIgnoreRowNum, varSheet=0):

        # 3.7 获取某些列的列数据(可忽略多行)
        # print(Openpyxl_PO.getColValueByCol([1, 3], [1, 2]))  # 获取第二列和第四列的列值，并忽略第1，2行的行值。
        # print(Openpyxl_PO.getColValueByCol([2], [], "python"))  # 获取第2列所有值。
        l_colData = []  # 每列的数据
        l_allData = []  # 所有的数据
        try:
            sh = self.sh(varSheet)
        except:
            print("[Error], " + varSheet + "不存在！")
            sys.exit(0)
        for col in l_varCol:
            for row in range(1, sh.max_row+1):
                if row not in l_varIgnoreRowNum:
                    l_colData.append(sh.cell(row, col).value)
            l_allData.append(l_colData)
            l_colData = []
        return l_allData


    def getCoordinate(self, varRow, varCol, varSheet=0):

        # 3.8 获取单元格的坐标
        sh = self.sh(varSheet)
        return sh.cell(row=varRow, column=varCol).coordinate


    def getDimensions(self, varSheet=0):

        # 3.9 获取所有数据的坐标
        sh = self.sh(varSheet)
        return (sh.dimensions)



    # todo [清除]

    def clsRow(self, varNums, varSheet=0):

        # 4.1 清空行
        # Openpyxl_PO.clsRow(2)  # 清空第2行
        sh = self.sh(varSheet)
        for i in range(1, sh.max_row):
            sh.cell(row=varNums, column=i, value="")
        self.save()


    def clsCol(self, varCol, varSheet=0):

        # 4.2 清空列
        # Openpyxl_PO.clsCol(2)  # 清空第2列
        sh = self.sh(varSheet)
        for i in range(1, sh.max_row+1):
            sh.cell(row=i, column=varCol).value = None
        self.save()


    def clsColRetainTitle(self, varCol, varSheet=0):

        # 4.2.1 清空列保留标题
        sh = self.sh(varSheet)
        for i in range(sh.max_row-1):
            sh.cell(row=i + 2, column=varCol).value = None
        self.save()


    def delSeriesRow(self, varFrom, varSeries=1, varSheet=0):

        # 4.3 删除连续行
        # Openpyxl_PO.delSeriesRow(2, 3)  # 删除从第二行开始连续三行数据 （即删除2，3，4行）
        sh = self.sh(varSheet)
        sh.delete_rows(idx=varFrom, amount=varSeries)
        self.save()



    def delSeriesCol(self, varFrom, varSeries=1, varSheet=0):

        # 4.4 删除连续列
        # Openpyxl_PO.delSeriesCol(2, 3)  # 删除从第二列开始连续三列数据 （即删除2，3，4列）
        # Openpyxl_PO.delSeriesCol('U', 1)  # 删除从第U列开始连续1列数据 （即删除U列）
        sh = self.sh(varSheet)
        if isinstance(varFrom, int):
            sh.delete_cols(idx=varFrom, amount=varSeries)
        else:
            sh.delete_cols(idx=column_index_from_string(varFrom), amount=varSeries)
        self.save()

    # todo [多表]

    def getDiffValueByCmp(self, l_file1row, l_file2row):

        '''
        5.2 两工作表比较，对差异内容标注颜色
        :param l_file1row:
        :param l_file2row:
        :return:
            print(Openpyxl_PO.getDiffValueByLeft(Openpyxl_PO.getRowValue(), Openpyxl_PO2.getRowValue()))
            [[5, 'member_id', 1311441], [7, 'loan_amnt', 5600]]   表示 第五行，member_id列的值1311441
            [[5, 'member_id', 5555], [7, 'loan_amnt', 1200]]
        '''

        d_left = {}
        d_left_sub = {}
        d_right = {}
        d_right_sub = {}
        d_all = {}

        if len(l_file1row) == len(l_file2row):
            for i in range(len(l_file1row)):
                if l_file1row[i] != l_file2row[i]:
                    for j in range(len(l_file1row[i])):
                        if l_file1row[i][j] != l_file2row[i][j]:
                            d_left_sub[l_file1row[0][j]] = l_file1row[i][j]
                            d_right_sub[l_file2row[0][j]] = l_file2row[i][j]
                    d_left[i + 1] = d_left_sub
                    d_right[i + 1] = d_right_sub
                    d_left_sub = {}
                    d_right_sub = {}

            if d_left != {} or d_right != {}:
                d_all["left"] = d_left
                d_all["right"] = d_right
                return d_all
            else:
                print("[ok], 两列表比对结果一致")
        else:
            print("[warning], 两列表数量不一致！")


    def setColorByDiff(self, varSheet1, varSheet2):

        # 5.2 两工作表比较，对差异内容标注颜色
        # 前提条件，两sheet表的行列数一致
        # Openpyxl_PO.setColorByDiff("Sheet1", "Sheet2")

        l_sheetOneRow = self.getRowValue(varSheet1)
        l_sheetTwoRow = self.getRowValue(varSheet2)

        if l_sheetOneRow == None or l_sheetTwoRow == None:
            print("[Error], " + varSheet1 + " 或 " + varSheet2 +" 不存在！")
            sys.exit(0)

        if len(l_sheetOneRow) == len(l_sheetTwoRow):
            for i in range(len(l_sheetOneRow)):
                for j in range(len(l_sheetOneRow[i])):
                    if l_sheetOneRow[i][j] != l_sheetTwoRow[i][j]:
                        self.setCellColor(i+1, j+1, "FF0000", varSheet1)
                        self.setCellColor(i+1, j+1, "ffeb9c", varSheet2)
                print("检查第" + str(i+1) + "行")

            self.save()
        else:
            print("[warning], 两sheet的行数不一致！")
            sys.exit(0)


    def setSheetByDiff(self, varSheet1, varSheet2):

        # 5.3 两工作表比较，生成新表Sheet1%Sheet2，对差异内容标注颜色
        # 前提条件，两sheet表的行列数一致
        # Openpyxl_PO.setSheetByDiff("Sheet1", "Sheet2")

        l_sheetOneRow = self.getRowValue(varSheet1)
        l_sheetTwoRow = self.getRowValue(varSheet2)

        if l_sheetOneRow == None or l_sheetTwoRow == None:
            print("[Error], " + varSheet1 + " 或 " + varSheet2 +" 不存在！")
            sys.exit(0)

        if len(l_sheetOneRow) == len(l_sheetTwoRow):

            # 生成临时sheet
            varSheet = varSheet1 + "%" + varSheet2
            self.delSheet(varSheet)
            self.addSheetCover(varSheet, 99)

            for i in range(len(l_sheetOneRow)):
                for j in range(len(l_sheetOneRow[i])):
                    if l_sheetOneRow[i][j] == "" and l_sheetTwoRow[i][j] == "":
                        pass
                    elif l_sheetOneRow[i][j] != l_sheetTwoRow[i][j]:
                        print(l_sheetOneRow[i][j], l_sheetTwoRow[i][j])
                        self.setCellValue(i+1, j+1, str(l_sheetOneRow[i][j]) + "/" + str(l_sheetTwoRow[i][j]), varSheet)
                        self.setCellColor(i+1, j+1, "FF0000", varSheet)
                    else:
                        self.setCellValue(i+1, j+1, str(l_sheetOneRow[i][j]), varSheet)
            self.save()
            return varSheet
        else:
            print("[warning], 两sheet的行数不一致！")
            sys.exit(0)


    def moveValue(self, varFrom, varRows, varCols, varSheet=0):

        # 6 移动范围数据
        sh = self.sh(varSheet)
        sh.move_range(varFrom, rows=varRows, cols=varCols)
        self.save()



if __name__ == "__main__":

    Sys_PO.killPid('EXCEL.EXE')
    Openpyxl_PO = OpenpyxlPO("ExcelPO/i_erp_reportField_case.xlsx")


    # print("1.1 新建".center(100, "-"))
    # Openpyxl_PO.newExcel("./OpenpyxlPO/newfile2.xlsx", "mySheet1", "mySheet2", "mySheet3")  # 新建excel，生成三个工作表（mySheet1,mySheet2,mySheet3），默认定位在第一个mySheet1表。

    # print("1.6 添加工作表(不覆盖)".center(100, "-"))
    # Openpyxl_PO.addSheet("saasuser1")

    # print("1.7 添加工作表(覆盖)".center(100, "-"))
    # Openpyxl_PO.addSheetCover("Sheet1", 1)    # 当index足够大时，则在最后一个位置添加工作表
    # Openpyxl_PO.open()

    # print("1.8 删除工作表".center(100, "-"))
    # Openpyxl_PO.delSheet("Sheet1")
    # Openpyxl_PO.delSheet("mySheet1")

    # print("2.0.1 插入一行或多行".center(100, "-"))
    # Openpyxl_PO.insertRows(3)   # 删除第三行到最后行的内容
    # Openpyxl_PO.insertRows(3, 1)  # 在第三行前插入一行空白
    # Openpyxl_PO.insertRows(3, 5)  # 在第三行前插入五行空白
    # print("2.0.2 插入一列或多列".center(100, "-"))
    # Openpyxl_PO.insertCols(5)  # 删除第五列到最后列的内容
    # Openpyxl_PO.insertCols(3, 1)  # 在第三列前插入一列空白



    # print("2.1 设置单元格值".center(100, "-"))
    # Openpyxl_PO.setCellValue(1, 6, "jinhao")

    # print("2.2 设置整行值".center(100, "-"))
    # Openpyxl_PO.setRowValue({7: ["你好", 12345, "7777"], 8: ["44", None, "777777777"]})  # 对最后一个sheet表，对第7，8行分别写入内容，如遇None则跳过该单元格
    # Openpyxl_PO.setRowValue({7: ["你好", 12345, "7777"], 8: ["44", "None", "777777777"]}, -1)  # 对最后一个sheet表，对第7，8行分别写入内容
    # Openpyxl_PO.open()

    # print("2.3 设置整列值".center(100, "-"))
    # Openpyxl_PO.setColValue({"A": ["k1", 666, "777"], "F": ["4456", None, "888"]}, -1)  # 对最后一个sheet表，对第7，8行分别写入内容，如遇None则跳过该单元格
    # Openpyxl_PO.open()

    # print("2.4 追加整行值".center(100, "-"))
    # Openpyxl_PO.addOnRowValue([['姓名', '电话', '成绩', '学科'], ['毛泽东', 15266606298, 14, '化学'], ['周恩来', 15201077791, 78, '美术']])   # 保留原有数据，在原数据下生成数据。

    # print("2.5 设置单元格行高与列宽".center(100, "-"))
    # Openpyxl_PO.setCellDimensions(3, 30, 'f', 34)  # 设置第三行行高30，第f列宽34

    # print("2.5.2 设置单行多列行高与列宽".center(100, "-"))
    # Openpyxl_PO.setRowColDimensions(5, 30, ['f', 'h'], 33)  # 设置第五行行高30，f - h列宽33

    # print("2.6 设置所有单元格的行高与列宽".center(100, "-"))
    # Openpyxl_PO.setAllCellDimensions(30, 20)

    # print("2.7 设置所有单元格自动换行".center(100, "-"))
    # Openpyxl_PO.setAllWordWrap()
    # Openpyxl_PO.setAllWordWrap("Sheet1")

    # print("2.8 设置冻结首行".center(100, "-"))
    # Openpyxl_PO.setFreeze('A2', "saasuser")



    # print("2.9 设置单元格对齐样式".center(100, "-"))
    # Openpyxl_PO.setCellAlignment(5, 4, 'center', 'top')
    # Openpyxl_PO.setCellAlignment(1, "e", 'center', 'top', 45)
    # Openpyxl_PO.setCellAlignment(1, 1, 'center', 'top', 45, True)
    # Openpyxl_PO.setCellAlignment(5, 4, 'center', 'center', "saasuser1")

    # print("2.9.2 设置单行多列对齐样式".center(100, "-"))
    # Openpyxl_PO.setRowColAlignment(1, ["c", "e"], 'center', 'center')  # 第一行第c,d,e列居中
    # Openpyxl_PO.setRowColAlignment(9, "all", 'center', 'center')  # 第九行全部居中

    # print("2.9.3 设置所有单元格对齐样式".center(100, "-"))
    # Openpyxl_PO.setAllCellAlignment('center', 'center')


    # print("2.10 设置筛选列".center(100, "-"))
    # Openpyxl_PO.setFilterCol("all")  # 全部筛选
    # Openpyxl_PO.setFilterCol("") # 取消筛选
    # Openpyxl_PO.setFilterCol("A2") # 对A2筛选



    # print("2.11 设置单元格字体（字体、字号、粗斜体、下划线、颜色）".center(100, "-"))
    # Openpyxl_PO.setCellFont(1, 6)  # 设置第一行第六列字体（默认微软雅黑字号16粗体）
    # Openpyxl_PO.setCellFont(2, "f")  # 设置第一行第f列字体（默认微软雅黑字号16粗体）
    # Openpyxl_PO.setCellFont(5, "f", size=14, bold=True, color="ff0000")
    # Openpyxl_PO.setCellFont(5, "f", size=14, bold=True)

    # print("2.11.2 设置单行多列字体".center(100, "-"))
    # Openpyxl_PO.setRowColFont(1, ["b", "h"])  # 第一行第b-h列
    # Openpyxl_PO.setRowColFont(9, "all")  # 第九行

    # print("2.11.3 设置所有单元格字体".center(100, "-"))
    # Openpyxl_PO.setAllCellFont()




    # print("2.12 设置单元格边框".center(100, "-"))
    # Openpyxl_PO.setBorder(1, 2, left = ['thin','ff0000'], right = ['thick','ff0000'], top = ['thin','ff0000'],bottom = ['thick','ff0000'])

    # print("2.13 设置单元格填充背景色".center(100, "-"))
    # Openpyxl_PO.setPatternFill(2, 2, 'solid', '006100')  # 单元格背景色

    # print("2.14 设置单元格填充渐变色".center(100, "-"))
    # Openpyxl_PO.setGradientFill(3, 3, stop=["FFFFFF", "99ccff", "000000"])



    # print("2.15 设置单元格背景色".center(100, "-"))
    # Openpyxl_PO.setCellColor(5, 1)  # 清除第5行第1列的背景色
    # Openpyxl_PO.setCellColor(5, "d")  # 清除第5行d列的背景色
    # Openpyxl_PO.setCellColor(5, 1, "ff0000", "Sheet2")  # 设置第五行第1列设置红色
    # Openpyxl_PO.setCellColor(5, "e", "ff0000")  # 设置第五行e列设置红色
    # Openpyxl_PO.setCellColor(None, None)  # 清除所有背景色

    # print("2.15.2 设置单行多列背景色".center(100, "-"))
    # Openpyxl_PO.setRowColColor(5, ['b', 'd'], "ff0000") # 设置第五行第b，c，d列背景色
    # Openpyxl_PO.setRowColColor(7, "all", "ff0000")  # 设置第五行所有列背景色

    # print("2.15.3 设置所有单元格背景色".center(100, "-"))
    # Openpyxl_PO.setAllCellColor("ff0000")  # 设置所有单元格背景色
    # Openpyxl_PO.setAllCellColor(None)  # 清除所有单元格背景色



    # print("2.16 设置整行(可间隔)背景色".center(100, "-"))
    # Openpyxl_PO.setRowColor(5, 0, "ff0000")  # 从第3行开始每行颜色标红色
    # Openpyxl_PO.setRowColor(3, 1, "ff0000")  # 从第3行开始每隔1行颜色标红色

    # print("2.17 设置整列(可间隔)背景色".center(100, "-"))
    # Openpyxl_PO.setColColor(2, 0, "ff0000")  # 从第2列开始每列颜色为红色
    # Openpyxl_PO.setColColor(2, 1, "ff0000")  # 从第2列开始每隔1列设置颜色为红色

    # print("2.18 设置工作表背景颜色".center(100, "-"))
    # Openpyxl_PO.setSheetColor("FF0000")




    # print("3.1 获取总行数和总列数".center(100, "-"))
    # print(Openpyxl_PO.getRowCol())  # [4,3] //返回第1个工作表的总行数和总列数
    # print(Openpyxl_PO.getRowCol(1))  # [4,3] //返回第2个工作表的总行数和总列数
    # print(Openpyxl_PO.getRowCol("python"))  # [4,3] //返回python工作表的总行数和总列数

    # print("3.2 获取单元格值".center(100, "-"))
    # print(Openpyxl_PO.getCellValue(3, 2))  # 获取第3行第2列的值

    # print("3.3 获取单行数据".center(100, "-"))
    # print(Openpyxl_PO.getOneRowValue(2))

    # print("3.4 获取每行数据".center(100, "-"))
    # print(Openpyxl_PO.getRowValue())
    # print(Openpyxl_PO.getRowValue("browser%interface"))
    #
    # print("3.5 获取每列数据".center(100, "-"))
    # print(Openpyxl_PO.getColValue())
    # # print(Openpyxl_PO.getColValue("python"))
    #
    # print("3.6 获取指定列的行数据".center(100, "-"))
    # print(Openpyxl_PO.getRowValueByCol([1, 2, 4], -1))   # 获取最后一个工作表的第1，2，4列的行数据
    #
    # print("3.7 获取某些列的列数据(可忽略多行)".center(100, "-"))
    # print(Openpyxl_PO.getColValueByCol([1, 3], [1, 2]))   # 获取第二列和第四列的列值，并忽略第1，2行的行值。
    # print(Openpyxl_PO.getColValueByCol([2], [], "上海"))  # 获取第2列所有值。

    # print("3.8 获取单元格的坐标".center(100, "-"))
    # print(Openpyxl_PO.getCoordinate(2, 5))   # E2

    # print("3.9 获取工作表数据的坐标".center(100, "-"))
    # print(Openpyxl_PO.getDimensions())  # A1:E17



    # print("4.1 清空行".center(100, "-"))
    # Openpyxl_PO.clsRow(2)  # 清空第2行
    #
    # print("4.2 清空列".center(100, "-"))
    # Openpyxl_PO.clsCol(2)  # 清空第2列
    # print("4.2.1 清空列保留标题".center(100, "-"))
    # Openpyxl_PO.clsColRetainTitle(2)  # 清空第2列保留标题

    # print("4.3 删除行".center(100, "-"))
    # # Openpyxl_PO.delSeriesRow(2, 3)  # 删除第2行之连续3行（删除2，3，4行）
    #
    # print("4.4 删除列".center(100, "-"))
    # # Openpyxl_PO.delSeriesCol(1, 2)  # 删除第1列之连续2列（删除1，2列）
    # # Openpyxl_PO.delSeriesCol(2, 1, "python")  # 删除第2列之连续1列（删除2列）
    # Openpyxl_PO.delSeriesCol('D', 1)  # 删除第D列之连续1列（删除D列）



    # print("5.1 两表比较获取差异内容（两表标题与行数必须一致） ".center(100, "-"))
    # Openpyxl_PO = OpenpyxlPO("./data/loanStats.xlsx")
    # Openpyxl_PO2 = OpenpyxlPO("./data/loanStats2.xlsx")
    # print(Openpyxl_PO.getDiffValueByCmp(Openpyxl_PO.getRowValue("Sheet2"), Openpyxl_PO2.getRowValue("Sheet2")))

    # # print("5.2 对一张表的两个sheet进行数据比对，差异数据标注颜色 ".center(100, "-"))
    # Openpyxl_PO = OpenpyxlPO("./data/loanStats.xlsx")
    # Openpyxl_PO.setColorByDiff("Sheet1", "Sheet2")

    # # print("5.3 对一张表的两个sheet进行数据比对，将结果写入第一个sheet ".center(100, "-"))
    # Openpyxl_PO.setSheetByDiff("browser", "interface")


    # # print("6 移动范围数据".center(100, "-"))
    # Openpyxl_PO.moveValue('C1:D2', 3, -2)  # 把'C1:D2'区域移动到 下三行左二列
    # Openpyxl_PO.moveValue('A1:C14', 0, 3)  # 把'A1:C14'区域向右移动3列


    r = Openpyxl_PO.getRowCol("browser%interface")[0]
    c = Openpyxl_PO.getRowCol("browser%interface")[1]

    varSign = 0
    list11 = []
    for i in range(r):
        for j in range(c):
            if "/" in Openpyxl_PO.getCellValue(i + 1, j + 1, "browser%interface"):
                varSign = 1
        if varSign == 1:
            list11.append("error")
        else:
            list11.append("ok")
        varSign = 0

    Openpyxl_PO.insertCols(1, 1, "browser%interface")
    Openpyxl_PO.setColValue({"A": list11},"browser%interface")
    Openpyxl_PO.setCellValue(1,1,"result","browser%interface")


    Openpyxl_PO.open()

