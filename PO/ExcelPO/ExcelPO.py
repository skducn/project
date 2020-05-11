# -*- coding: utf-8 -*-
# *********************************************************************
# Author        : John
# Date          : 2019-9-26
# Description   : excel 对象层
# http://www.cnblogs.com/snow-backup/p/4021554.html
# python处理Excel，pythonexcel http://www.bkjia.com/Pythonjc/926154.html
# Python利用pandas处理Excel数据的应用 # https://www.cnblogs.com/liulinghua90/p/9935642.html

# python中处理excel表格，常用的库有xlrd（读excel）表、xlwt（写excel）表、openpyxl（可读写excel表）等。xlrd读数据较大的excel表时效率高于openpyxl，建议使用 xlrd 和 xlwt 两个库。
# openpyxl , xlsxwriter xlrd xlwt xlutils 下载地址：http://www.python-excel.org/
# 以上这些库都没有提供修改 excel 表格内容功能，一般只能将原excel中的内容读出、做完处理后，再写入一个新的excel文件。
#
# 常见问题1：处理 excel 表格时遇到的 unicode编码。
# python默认字符编码为unicode，所以从excel中读取的中文sheet或中文名时，会提示“UnicodeEncodeError: 'ascii' codec can't encode characters in position 0-2: ordinal not in range(128)。”
# 这是由于 windows里中文使用了gb2312编码方式，python误将其当作 unicode 和 ascii 来解码所造成的错误，使用“.encode('gb2312')”即可解决打印中文的问题。
# 可在文件名前加‘u’表示将该中文文件名采用unicode编码。

# 问题2：处理excel表格时遇到 excel中 时间显示错误。
# 有excel中，时间和日期都使用浮点数表示。如单元格中时间是 ‘2013年3月20日’，python输出为‘41353.0’；当其单元格格式改变为日期后，内容又变为了‘2013年3月20日’。
# 而使用xlrd读出excel中的日期和时间后，得到是的一个浮点数。所以当向excel中写入的日期和时间为一个浮点数也不要紧，只需将表格的表示方式改为日期和时间，即可得到正常的表示方式。
# excel中，用浮点数1表示1899年12月31日。

# python读取excel中单元格的内容返回的有5种类型，即上面例子中的ctype:
# ctype： 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
# 如 sh.cell(2, 1).ctype == 3  判断某一单元格内容是否是日期
# *********************************************************************

from datetime import date
import xlrd, xlwt
from xlutils.copy import copy
from openpyxl import load_workbook
import openpyxl, os, sys
# import pandas as pd


'''
1，新建excel（by openpyxl）
2.1，单个元素写操作（by openpyxl）
2.2，多个元素写操作（by openpyxl）
2.3，单个元素及批量行列写操作（by xlrd）

3.1，获取所有工作表名
3.2，获取单个工作表名
3.4，获取行列数
3.5，获取行值
3.6，获取列值

4，设置表格样式

5.1，删除行列
5.2，清空行列

6，两表比较，输出diff
'''

class ExcelPO():

    def __init__(self):
        pass

    # 1，新建excel（by openpyxl）
    def createExcel(self, varFileName, *varSheetName):
        # 新建excel，生成N个工作表（默认一个Sheet1）
        # 注意：如果文件已存在则先删除再新建
        # Excel_PO.createExcel("d:\\test1.xlsx")  # 新建excel，默认生成一个工作表Sheet1
        # Excel_PO.createExcel("d:\\test1.xlsx", "mySheet1")  # 新建excel，生成一个工作表mySheet1
        # Excel_PO.createExcel("d:\\test1.xlsx", "mySheet1", "mySheet2", "mySheet3")  # 新建excel，生成三个工作表（mySheet1,mySheet2,mySheet3），默认定位在第一个mySheet1表。
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            if len(varSheetName) == 0:
                ws.title = "Sheet1"
            else:
                ws.title = varSheetName[0]
            for i in range(1, len(varSheetName)):
                wb.create_sheet(varSheetName[i])
            wb.save(varFileName)
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")


    # 2.1，单个元素写操作（by openpyxl）
    def writeXlsx(self, varFileName, varSheet, varRow, varCol, varContent):
        # 对单元格进行写操作（只支持.xlsx）
        # Excel_PO.writeXlsx("excel3.xlsx", "Sheet1", 5, 3, "测试一下")  # 第五行第三列写入内容
        if os.path.isfile(varFileName) == False:
            self.createExcel(varFileName, varSheet)
        try:
            wb = load_workbook(varFileName)
            xl_sheet_names = wb.sheetnames  # 获取所有sheet页名字
            # 判断 varSheet 是数字还是字符
            if isinstance(varSheet, int):  # 判断是int类型
                # 定位到sheet页,[0]为sheet页索引
                wk_sheet = wb[xl_sheet_names[varSheet]]
            elif isinstance(varSheet, str):
                wk_sheet = wb[varSheet]
            else:
                print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")
            wk_sheet.cell(row=varRow, column=varCol, value=varContent)
            wb.save(varFileName)
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 2.2，多个元素写操作（by openpyxl）
    def writeXlsxByMore(self, varFileName, varSheet, varList_Row_Col_Content):
        # 对多个元素写操作（只支持.xlsx）
        # Excel_PO.writeXlsxByMore("d:\\test3.xlsx", "mySheet1", [[1, "你好", "测试", "报告"], [2, "再见", "", "好了"]])
        # 第一行写入 "你好", "测试", "报告"
        # 第二行写入 "再见", "", "好了"
        if os.path.isfile(varFileName) == False:
            self.createExcel(varFileName, varSheet)
        try:
            wb = load_workbook(varFileName)
            xl_sheet_names = wb.sheetnames  # 获取所有sheet页名字
            # 判断 varSheet 是数字还是字符
            if isinstance(varSheet, int):  # 判断是int类型
                # 定位到sheet页,[0]为sheet页索引
                wk_sheet = wb[xl_sheet_names[varSheet]]
            elif isinstance(varSheet, str):
                wk_sheet = wb[varSheet]
            else:
                print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")
            for i in range(len(varList_Row_Col_Content)):
                for j in range(1, len(varList_Row_Col_Content[i])):
                    wk_sheet.cell(row=varList_Row_Col_Content[i][0], column=j, value=varList_Row_Col_Content[i][j])
            wb.save(varFileName)
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 2.3，单个元素及批量行列写操作（by xlrd）
    def writeXls(self, varFileName, varSheet, varRow, varCol, varContent, varIgnore=""):
        # 功能：写表操作，对单元格进行单个及批量写入（只支持.xls）
        # Excel_PO.writeXls("excel1.xls", "Sheet2", 2, 5, "令狐冲")  # 对Sheet2，第2行第5列写入令狐冲
        # Excel_PO.writeXls("excel1.xls", "Sheet2", "*", 5, "")  # 对Sheet2，清除第6列 (保留第一行，一般第一行是标题)
        # Excel_PO.writeXls("excel1.xls", 1, 2, '*', 'aaa')  # 对Sheet2，第3行写入aaa (保留第一列，一般第一列是标题)
        try:
            wb = xlrd.open_workbook(filename=varFileName)
            if isinstance(varSheet, int):  # 判断是int类型
                sh = wb.sheet_by_index(varSheet)
            elif isinstance(varSheet, str):
                sh = wb.sheet_by_name(varSheet)
            else:
                print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")
            wbk = copy(wb)
            sheet = wbk.get_sheet(varSheet)
            if varRow == "*":
                for i in range(sh.nrows-varIgnore):
                    sheet.write(i+varIgnore, varCol-1, varContent)
            elif varCol == '*':
                for i in range(sh.ncols-varIgnore):
                    sheet.write(varRow-1, i+varIgnore, varContent)
            else:
                sheet.write(varRow-1, varCol-1, varContent)
            wbk.save(varFileName)
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")


    # 3.1，获取所有工作表名
    def getAllSheetName(self, varFileName):
        try:
            wb = xlrd.open_workbook(filename=varFileName)
            return wb.sheet_names()
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 3.2，获取单个工作表名
    def getSheetNameByIndex(self, varFileName, varIndex):
        # 通过 index 获取工作表名，0=第一个Sheet，1=第二个Sheet，-1表示从后往前数，以此类推
        # print(Excel_PO.getSheetNameByIndex("d:\\test3.xlsx", 0))  # mySheet1
        # print(Excel_PO.getSheetNameByIndex("d:\\test3.xlsx", 1))  # mySheet2
        # print(Excel_PO.getSheetNameByIndex("d:\\test3.xlsx", -1))  # mySheet3
        try:
            wb = xlrd.open_workbook(filename=varFileName)
            return wb.sheet_by_index(varIndex).name
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 3.3，获取行列数
    def getRowCol(self, varFileName, varSheet=0):
        # 获取行数和列数
        # print(Excel_PO.getRowCol("d:\\test13.xlsx"))  # [5, 3]  // 默认定位第一张工作表，返回列表[行，列]
        # print(Excel_PO.getRowCol("d:\\test13.xlsx")[0])  # 5 // 默认定位第一张工作表，返回行
        # print(Excel_PO.getRowCol("d:\\test13.xlsx")[1])  # 3  // 默认定位第一张工作表，返回列
        # print(Excel_PO.getRowCol("d:\\test3.xlsx", "mySheet1"))  # [10, 4] //返回mySheet1工作表的行和列
        # print(Excel_PO.getRowCol("d:\\test3.xlsx", 1))  # [11, 12]  // 定位第二张工作表，返回行列列表
        list1 = []
        try:
            wb = xlrd.open_workbook(varFileName)
            if isinstance(varSheet, int):
                sh = wb.sheet_by_index(varSheet)
            else:
                sh = wb.sheet_by_name(varSheet)
            if wb.sheet_loaded(varSheet) == True:  # 检查 sheet是否导入完毕，返回True 或 False
                cols = sh.ncols
                rows = sh.nrows
                list1.append(rows)
                list1.append(cols)
            return (list1)
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 3.4，获取行值
    def getRowValue(self, varFileName, varRow, varSheet=0):
        # 获取行值，以列表形式返回，如 ['小寒1', '女', 58665985.0, datetime.date(2013, 3, 20), 134.0] , 异常返回 None，默认sheetNO=0定位第一个工作簿。
        # 对日期时间进行了处理，可正常返回 2013-03-20
        # 可使用切片获取某行某个单元值
        # print(Excel_PO.readRowValue("excel1.xlsx", 2, 1))  # ['小寒1', '女', 58665985.0, datetime.date(2013, 3, 20), 134.0]
        # print(Excel_PO.readRowValue("excel1.xlsx", 2, "Sheet2"))  # ['小寒1', '女', 58665985.0, datetime.date(2013, 3, 20), 134.0]
        # print(Excel_PO.readRowValue("excel1.xls", 1)[3])  # 希缇缇  使用切片获取某行某个单元值
        list1 = []
        wb = xlrd.open_workbook(varFileName)
        if isinstance(varSheet, int):  # 判断变量的类型，int list tuple dict str 参考 https://www.cnblogs.com/fmgao-technology/p/9065753.html
            sh = wb.sheet_by_index(varSheet)
        else:
            sh = wb.sheet_by_name(varSheet)
        cols = sh.ncols
        try:
            for r in range(cols):
                if sh.cell(varRow - 1, r).ctype == 3:
                    # value = xldate_as_datetime(sh.cell_value(rowx=varRow, colx=r), 0)
                    value = xlrd.xldate_as_tuple(sh.cell_value(rowx=varRow-1, colx=r-1), wb.datemode)
                    value = date(*value[:3])
                else:
                    value = sh.cell_value(rowx=varRow-1, colx=r)
                list1.append(value)
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")
        return list1

    # 3.5，获取列值
    def getColValue(self, varFileName, varCol, varSheet=0):
        # 获取 某一列的值，以列表形式返回，如 ['赵云', '小寒', '陈晓东'] , 异常或为空则返回 None
        try:
            list1 = []
            wb = xlrd.open_workbook(varFileName)
            if isinstance(varSheet,
                          int):  # 判断变量的类型，int list tuple dict str 参考 https://www.cnblogs.com/fmgao-technology/p/9065753.html
                sh = wb.sheet_by_index(varSheet)
            else:
                sh = wb.sheet_by_name(varSheet)
            rows = sh.nrows
            try:
                for r in range(rows):
                    if sh.cell(r, varCol).ctype == 3:
                        value = xlrd.xldate_as_tuple(sh.cell_value(rowx=r, colx=varCol), wb.datemode)
                        value = date(*value[:3])
                    else:
                        value = sh.cell_value(r, varCol-1)
                    list1.append(value)
            except:
                print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")
            return list1
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")


    # 4，设置表格样式
    def set_style(self, name, height, bold=False):
        style = xlwt.XFStyle()
        font = xlwt.Font()
        font.name = name
        font.bold = bold
        font.color_index = 4
        font.height = height
        style.font = font
        return style


    # 5.1，删除行列
    def delRowColForXlsx(self, varFileName, varSheet, varType, varFrom, varTo=1):
        # 功能删除多行或多列。（for xlsx）
        # Excel_PO.delRowColForXlsx("excel3.xlsx", "col", 1, 2)  # 从第1列开始，连续删除2列
        # Excel_PO.delRowColForXlsx("excel3.xlsx", "row", 2, 3)  # 从第2行开始，连续删除3行
        # wk = load_workbook(varFileName)
        # wk_name = wk.sheetnames
        # wk_sheet = wk[wk_name[0]]  # 默认第一张表
        try:
            wb = load_workbook(varFileName)  # 打开Excel
            xl_sheet_names = wb.sheetnames
            # print(xl_sheet_names)  # 打印所有sheet页名称
            # 判断 varSheet 是数字还是字符
            if isinstance(varSheet, int):  # 判断是int类型
                # 定位到相应sheet页,[0]为sheet页索引
                wk_sheet = wb[xl_sheet_names[varSheet]]
            elif isinstance(varSheet, str):
                wk_sheet = wb[varSheet]
            else:
                print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")
            if varType == "row":
                wk_sheet.delete_rows(varFrom, varTo)  # 删除从第一行开始算的2行内容
            elif varType == "col":
                wk_sheet.delete_cols(varFrom, varTo)  # 删除从第一列开始算的2列内容
            wb.save(varFileName)
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 5.2，清空单元格 for xlsx
    def clearRowColForXlsx(self, varFileName, varSheet, varType, varNums):
        # Excel_PO.clearRowColForXlsx("d:\\test3.xlsx", "mySheet1", "row", 2)  # 打开mySheet1，清空第2行。
        # Excel_PO.clearRowColForXlsx("d:\\test3.xlsx", 0, "col", 1)  # 打开mySheet1，清空第1列
        try:
            wb = load_workbook(varFileName)  # 打开Excel
            xl_sheet_names = wb.sheetnames  # 获取所有sheet页名字
            # 判断 varSheet 是数字还是字符
            if isinstance(varSheet, int):  # 判断是int类型
                # 定位到相应sheet页,[0]为sheet页索引
                wk_sheet = wb[xl_sheet_names[varSheet]]
            elif isinstance(varSheet, str):
                wk_sheet = wb[varSheet]
            else:
                print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")
            if varType == "col":
                for i in range(wk_sheet.max_row):
                    wk_sheet.cell(row=i+1, column=varNums, value="")  # 清除第row行的第col列
            elif varType == "row":
                for i in range(wk_sheet.max_row):
                    wk_sheet.cell(row=varNums, column=i+1, value="")  # 清除第row行的第col列
            wb.save(varFileName)
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")


    # 6，两表比较，输出差异
    def cmpExcel(self,varFileName1, varSheetName1, varFileName2, varSheetName2):
        # 6，两表比较，输出差异

        try:
            list1 = self.getRowCol(varFileName1, varSheetName1)
            list2 = self.getRowCol(varFileName2, varSheetName2)
            tmpList1 = []
            tmpList2 = []
            mainList1 = []
            mainList2 = []
            wb = load_workbook(varFileName1)
            wk_sheet = wb[varSheetName1]
            for i in range(1, list1[0]+1):
                tmpList1.append(i)
                for j in range(1, list1[1]+1):
                    tmpList1.append(wk_sheet.cell(row=i,column=j).value)
                mainList1.append(tmpList1)
                tmpList1 = []
            # print(mainList1)

            wb = load_workbook(varFileName2)
            wk_sheet = wb[varSheetName2]
            for i in range(1, list2[0] + 1):
                tmpList2.append(i)
                for j in range(1, list1[1] + 1):
                    tmpList2.append(wk_sheet.cell(row=i, column=j).value)
                mainList2.append(tmpList2)
                tmpList2 = []
            # print(mainList2)

            a = [x for x in mainList1 if x in mainList2]  # 两个列表中都存在
            b = [y for y in (mainList1) if y not in a]  # 两个列表中的不同元素，输出 mainList1 中差异部分
            c = [y for y in (mainList2) if y not in a]  # 两个列表中的不同元素，输出 mainList2 中差异部分

            if b == []:
                print("ok，两表一致")
            else:
                print(varFileName1 + " => " + str(b))
                print(varFileName2 + " => " + str(c))

        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")




if __name__ == "__main__":

    Excel_PO = ExcelPO()


    # print("1，新建excel（by openpyxl）".center(100, "-"))
    # Excel_PO.createExcel("d:\\test1.xlsx")  # 新建excel，默认生成一个工作表Sheet1
    # Excel_PO.createExcel("d:\\test1.xlsx", "mySheet1")  # 新建excel，生成一个工作表mySheet1
    # Excel_PO.createExcel("d:\\test1.xlsx", "mySheet1", "mySheet2", "mySheet3")  # 新建excel，生成三个工作表（mySheet1,mySheet2,mySheet3），默认定位在第一个mySheet1表。

    # print("2.1，单个元素写操作（by openpyxl）".center(100, "-"))
    # Excel_PO.writeXlsx("d:\\test12313.xlsx", "mySheet3", 5, 3, "测试一下")  # 对 mySheet3 的第5行第3列写入内容
    # Excel_PO.writeXlsx("d:\\test3.xlsx", 1, 5, 3, "测试一下13")  # 对第二个工作表的第5行第3列写入内容

    # print("2.2，多个元素写操作（by openpyxl）".center(100, "-"))
    # Excel_PO.writeXlsxByMore("d:\\test3.xlsx", "mySheet1", [[1,"你好","测试","报告"], [2,"再见", "", "好了"]])
    # listSub = []
    # listSub.append(1)
    # listSub.append("表名")
    # listSub.append("表说明")
    # listSub.append("字段名")
    # listSub.append("字段说明")
    # listSub.append("字段类型")
    # listSub.append("是否为空")
    # listMain = []
    # listMain.append(listSub)
    # print(listMain)
    # Excel_PO.writeXlsxByMore("d:\\test3.xlsx", "mySheet1", listMain)

    # print("2.3，单个元素及批量行列写操作（by xlrd）".center(100, "-"))
    # Excel_PO.writeXls("d:\\test.xls", "Sheet1", 2, 5, "令狐冲")  # 对 Sheet2 的第2行第5列写入令狐冲
    # Excel_PO.writeXls("d:\\test.xls", "Sheet1", "*", 5, "c", 2)  # 对 Sheet2 的第5列写入c，忽略前2行
    # Excel_PO.writeXls("d:\\test.xls", "Sheet1", 4, '*', 'b', 3)  # 对 Sheet2 的第4行写入b，忽略前3列


    # print("3.1，获取所有工作表名".center(100, "-"))
    # print(Excel_PO.getAllSheetName("d:\\test3.xlsx"))  # ['mySheet1', 'mySheet2', 'mySheet3']
    #
    # print("3.2，获取单个工作表名".center(100, "-"))
    # print(Excel_PO.getSheetNameByIndex("d:\\test3.xlsx", 0))  # mySheet1
    # print(Excel_PO.getSheetNameByIndex("d:\\test3.xlsx", 1))  # mySheet2
    # print(Excel_PO.getSheetNameByIndex("d:\\test3.xlsx", -1))  # mySheet3  //-1表示从后往前数
    # #
    # # print("3.3，获取行列数".center(100, "-"))
    # print(Excel_PO.getRowCol("d:\\test13.xlsx"))  # [5, 3]  // 默认定位第一张工作表，返回列表[行，列]
    # print(Excel_PO.getRowCol("d:\\test13.xlsx")[0])  # 5 // 默认定位第一张工作表，返回行
    # print(Excel_PO.getRowCol("d:\\test13.xlsx")[1])  # 3  // 默认定位第一张工作表，返回列
    # print(Excel_PO.getRowCol("d:\\test3.xlsx", "mySheet1"))  # [10, 4] //返回mySheet1工作表的行和列
    # print(Excel_PO.getRowCol("d:\\test3.xlsx", 1))  # [11, 12]  // 定位第二张工作表，返回行列列表
    #
    # print("3.4，获取行值".center(100, "-"))
    # print(Excel_PO.getRowValue("d:\\test3.xlsx", 1))  # [0.0, 0.0, 'yoyo', 43909.0]   //默认第一个工作表，获取第一行数据，注意日期变成了43909.0
    # print(Excel_PO.getRowValue("d:\\test3.xlsx", 2, "mySheet1"))  # [1.0, 1.0, '', '2020.3.19']   //打开mySheet1，获取第二行数据，注意数字是浮点数，如1.0
    # # print(Excel_PO.getRowValue("d:\\test3.xlsx", 1)[3])  # 43909.0  //定位第一张工作表，获取第一行第四列
    #
    # print("3.5，获取列值".center(100, "-"))
    # print(Excel_PO.getColValue("d:\\test3.xlsx", 1, "mySheet1"))  # [1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0] // 获取mySheet1表 第一列值
    # print(Excel_PO.getColValue("d:\\test3.xlsx", 1, 2))  # ['a', '', '', '', '', '']   //获取第三个工作表的第一列



    # print("5.1，删除行列".center(100, "-"))
    # Excel_PO.delRowColForXlsx("d:\\test3.xlsx", "mySheet1", "col", 2)  # 删除表mySheet1的第2列
    # Excel_PO.delRowColForXlsx("d:\\test3.xlsx", "mySheet1", "col", 3, 2)  # 删除表mySheet1的第3，4列
    # Excel_PO.delRowColForXlsx("d:\\test3.xlsx", 0, "row", 3, 2)  # 删除mySheet1的第3行,第4行

    # print("5.2，清空行列".center(100, "-"))
    # Excel_PO.clearRowColForXlsx("d:\\test3.xlsx", "mySheet1", "row", 2)  # 清空表mySheet1的第2行
    # Excel_PO.clearRowColForXlsx("d:\\test3.xlsx", 0, "col", 1)  # 清空表mySheet1的第1列


    # print("6，两表比较，输出差异".center(100, "-"))
    # Excel_PO.cmpExcel("d:\\file2.xlsx", "mySheet1", "d:\\file3.xlsx", "mySheet1")

