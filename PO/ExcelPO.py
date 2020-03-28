# -*- coding: utf-8 -*-
# *********************************************************************
# Author        : John
# Date          : 2019-9-26
# Description   : excel 对象层
# http://www.cnblogs.com/snow-backup/p/4021554.html
# python处理Excel，pythonexcel http://www.bkjia.com/Pythonjc/926154.html
# python中处理excel表格，常用的库有xlrd（读excel）表、xlwt（写excel）表、openpyxl（可读写excel表）等。xlrd读数据较大的excel表时效率高于openpyxl，建议使用 xlrd 和 xlwt 两个库。
# openpyxl , xlsxwriter xlrd xlwt xlutils 下载地址：http://www.python-excel.org/
# 以上这些库都没有提供修改 excel 表格内容功能，一般只能将原excel中的内容读出、做完处理后，再写入一个新的excel文件。
#
# 常见问题1：处理 excel 表格时遇到的 unicode编码。
# python默认字符编码为unicode，所以从excel中读取的中文sheet或中文名时，会提示“UnicodeEncodeError: 'ascii' codec can't encode characters in position 0-2: ordinal not in range(128)。”
# 这是由于 windows里中文使用了gb2312编码方式，python误将其当作 unicode 和 ascii 来解码所造成的错误，使用“.encode('gb2312')”即可解决打印中文的问题。
# 可在文件名前加‘u’表示将该中文文件名采用unicode编码。

# 问题2：处理excel表格时遇到 excel中 时间显示错误。
# 有excel中，时间和日期都使用浮点数表示。如单元格中时间是 ‘2013年3月20日’，python输出为‘41353.0’；当其单元格格式改变为日期后，内容又变为了‘2013年3月20日’。
# 而使用xlrd读出excel中的日期和时间后，得到是的一个浮点数。所以当向excel中写入的日期和时间为一个浮点数也不要紧，只需将表格的表示方式改为日期和时间，即可得到正常的表示方式。
# excel中，用浮点数1表示1899年12月31日。

# python读取excel中单元格的内容返回的有5种类型，即上面例子中的ctype:
# ctype： 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
# 如 sh.cell(2, 1).ctype == 3  判断某一单元格内容是否是日期
# *********************************************************************

from datetime import date
import xlrd,xlwt
from xlutils.copy import copy
from openpyxl  import load_workbook
# import pandas as pd
# https://www.cnblogs.com/liulinghua90/p/9935642.html

class ExcelPO():

    def __init__(self):
        pass

    # 获取工作表名列表
    def getAllSheetName(self, varFileName):
        # 获取excel的工作表名，以列表形式返回，如 ['Sheet1', 'Sheet2', 'Sheet3']
        wb = xlrd.open_workbook(filename=varFileName)
        return wb.sheet_names()

    # 获取单个工作表名
    def getSheetNameByIndex(self, varFileName, varIndex):
        # 通过index方式获取 excel的工作表名，以字符串形式返回，如 'Sheet1'
        wb = xlrd.open_workbook(filename=varFileName)
        return wb.sheet_by_index(varIndex).name

    # 获取总行数
    def getAllRows(self, varFileName, varSheet=0):
        # 获取总行数
        wb = xlrd.open_workbook(filename=varFileName)
        sh = wb.sheet_by_index(varSheet)
        return sh.nrows

    # 获取行列数，如[4,3]
    def getRowCol(self, varFileName, varSheet=0):
        # 获取行数和列数，以列表形式返回，如 [4,3]，表示 4行3列。可通过index或name方式定位工作表名
        list1 = []
        wb = xlrd.open_workbook(varFileName)
        if isinstance(varSheet, int):  # 判断变量的类型，int list tuple dict str 参考 https://www.cnblogs.com/fmgao-technology/p/9065753.html
            sh = wb.sheet_by_index(varSheet)
        else:
            sh = wb.sheet_by_name(varSheet)
        if wb.sheet_loaded(varSheet) == True:  # 检查 sheet是否导入完毕，返回True 或 False
            cols = sh.ncols
            rows = sh.nrows
            list1.append(rows)
            list1.append(cols)
        return (list1)

    # 获取行值
    def getRowValue(self, varFileName, varRow, varSheet=0):
        # 获取行值，以列表形式返回，如 ['小寒1', '女', 58665985.0, datetime.date(2013, 3, 20), 134.0] , 异常返回 None，默认sheetNO=0定位第一个工作簿。
        # 对日期时间进行了处理，可正常返回 2013-03-20
        # 可使用切片获取某行某个单元值
        # print(Excel_PO.readRowValue("excel1.xlsx", 2, 1))  # ['小寒1', '女', 58665985.0, datetime.date(2013, 3, 20), 134.0]
        # print(Excel_PO.readRowValue("excel1.xlsx", 2, "Sheet2"))  # ['小寒1', '女', 58665985.0, datetime.date(2013, 3, 20), 134.0]
        # print(Excel_PO.readRowValue("excel1.xls", 1)[3])  # 希缇缇  使用切片获取某行某个单元值
        list1 = []
        wb = xlrd.open_workbook(varFileName)
        if isinstance(varSheet,
                      int):  # 判断变量的类型，int list tuple dict str 参考 https://www.cnblogs.com/fmgao-technology/p/9065753.html
            sh = wb.sheet_by_index(varSheet)
        else:
            sh = wb.sheet_by_name(varSheet)
        cols = sh.ncols
        try:
            for r in range(cols):
                if sh.cell(varRow - 1, r).ctype == 3:
                    # value = xldate_as_datetime(sh.cell_value(rowx=varRow, colx=r), 0)
                    value = xlrd.xldate_as_tuple(sh.cell_value(rowx=varRow, colx=r), wb.datemode)
                    value = date(*value[:3])
                else:
                    value = sh.cell_value(rowx=varRow, colx=r)
                list1.append(value)
        except:
            return None
        return list1

    # 获取列值
    def getColValue(self, varFileName, varCol, varSheet=0):
        # 获取 某一列的值，以列表形式返回，如 ['赵云', '小寒', '陈晓东'] , 异常或为空则返回 None
        list1 = []
        wb = xlrd.open_workbook(varFileName)
        if isinstance(varSheet,
                      int):  # 判断变量的类型，int list tuple dict str 参考 https://www.cnblogs.com/fmgao-technology/p/9065753.html
            sh = wb.sheet_by_index(varSheet)
        else:
            sh = wb.sheet_by_name(varSheet)
        rows = sh.nrows
        try:
            for r in range(rows):
                if sh.cell(r, varCol).ctype == 3:
                    value = xlrd.xldate_as_tuple(sh.cell_value(rowx=r, colx=varCol), wb.datemode)
                    value = date(*value[:3])
                else:
                    value = sh.cell_value(r, varCol)
                list1.append(value)
        except:
            return None
        return list1

    # 设置表格样式
    def set_style(self, name, height, bold=False):
        style = xlwt.XFStyle()
        font = xlwt.Font()
        font.name = name
        font.bold = bold
        font.color_index = 4
        font.height = height
        style.font = font
        return style

    # 对 xls 写操作
    def writeXls(self, varFileName, varSheet, varRow, varCol, varContent):
        # 功能：写表操作，对单元格进行单个及批量写入（只支持.xls）
        # Excel_PO.writeXls("excel1.xls", "Sheet2", 2, 5, "令狐冲")  # 对Sheet2，第3行第6列写入令狐冲
        # Excel_PO.writeXls("excel1.xls", "Sheet2", "*", 5, "")  # 对Sheet2，清除第6列 (保留第一行，一般第一行是标题)
        # Excel_PO.writeXls("excel1.xls", 1, 2, '*', 'aaa')  # 对Sheet2，第3行写入aaa (保留第一列，一般第一列是标题)
        wb = xlrd.open_workbook(filename=varFileName)
        if isinstance(varSheet, int):  # 判断是int类型
            sh = wb.sheet_by_index(varSheet)
        elif isinstance(varSheet, str):
            sh = wb.sheet_by_name(varSheet)
        else:
            exit()
        wbk = copy(wb)
        sheet = wbk.get_sheet(varSheet)
        if varRow == "*":
            for i in range(sh.nrows):
                sheet.write(i+1, varCol, varContent)
        elif varCol == '*':
            for i in range(sh.nrows):
                sheet.write(varRow, i+1, varContent)
        else:
            sheet.write(varRow, varCol, varContent)
        wbk.save(varFileName)

    # 对 xlsx 写操作
    def writeXlsx(self, varFileName, varSheet, varRow, varCol, varContent):
        # 功能：写表操作，对单元格进行写操作（只支持.xlsx）
        # Excel_PO.writeXlsx("excel3.xlsx", "Sheet1", 5, 3, "测试一下")  # 第五行第三列写入内容

        wb = load_workbook(varFileName)  # 打开Excel
        xl_sheet_names = wb.get_sheet_names()  # 获取所有sheet页名字
        # print(xl_sheet_names)  # 打印所有sheet页名称
        # 判断 varSheet 是数字还是字符
        if isinstance(varSheet, int):  # 判断是int类型
            # 定位到相应sheet页,[0]为sheet页索引
            wk_sheet = wb.get_sheet_by_name(xl_sheet_names[varSheet])
        elif isinstance(varSheet, str):
            wk_sheet = wb.get_sheet_by_name(varSheet)
        else:
            exit()
        wk_sheet.cell(row=varRow, column=varCol, value=varContent)
        wb.save(varFileName)


    # 删除行列
    def delRowColForXlsx(self, varFileName, varSheet, varType, varFrom, varTo):
        # 功能删除多行或多列。（for xlsx）
        # Excel_PO.delRowColForXlsx("excel3.xlsx", "col", 1, 2)  # 从第1列开始，连续删除2列
        # Excel_PO.delRowColForXlsx("excel3.xlsx", "row", 2, 3)  # 从第2行开始，连续删除3行
        # wk = load_workbook(varFileName)
        # wk_name = wk.sheetnames
        # wk_sheet = wk[wk_name[0]]  # 默认第一张表

        wb = load_workbook(varFileName)  # 打开Excel
        xl_sheet_names = wb.get_sheet_names()  # 获取所有sheet页名字
        # print(xl_sheet_names)  # 打印所有sheet页名称
        # 判断 varSheet 是数字还是字符
        if isinstance(varSheet, int):  # 判断是int类型
            # 定位到相应sheet页,[0]为sheet页索引
            wk_sheet = wb.get_sheet_by_name(xl_sheet_names[varSheet])
        elif isinstance(varSheet, str):
            wk_sheet = wb.get_sheet_by_name(varSheet)
        else:
            exit()

        if varType == "row":
            wk_sheet.delete_rows(varFrom, varTo)  # 删除从第一行开始算的2行内容
        elif varType == "col":
            wk_sheet.delete_cols(varFrom, varTo)  # 删除从第一列开始算的2列内容
        wb.save(varFileName)

    # 清空单元格 for xlsx
    def clearRowColForXlsx(self, varFileName, varSheet, varType, varNums):
        # 清空整行整列。

        wb = load_workbook(varFileName)  # 打开Excel
        xl_sheet_names = wb.get_sheet_names()  # 获取所有sheet页名字
        # print(xl_sheet_names)  # 打印所有sheet页名称
        # 判断 varSheet 是数字还是字符
        if isinstance(varSheet, int):  # 判断是int类型
            # 定位到相应sheet页,[0]为sheet页索引
            wk_sheet = wb.get_sheet_by_name(xl_sheet_names[varSheet])
        elif isinstance(varSheet, str):
            wk_sheet = wb.get_sheet_by_name(varSheet)
        else:
            exit()

        if varType == "col":
            for i in range(wk_sheet.max_row):
                wk_sheet.cell(row=i+1, column=varNums, value="")  # 清除第row行的第col列
        elif varType == "row":
            for i in range(wk_sheet.max_row):
                wk_sheet.cell(row=varNums, column=i+1, value="")  # 清除第row行的第col列
        wb.save(varFileName)


if __name__ == "__main__":
    Excel_PO = ExcelPO()
    # print(Excel_PO.getAllSheetName("excel3.xlsx"))  # ['Sheet1', 'Sheet2', 'Sheet3']
    # print(Excel_PO.getSheetNameByIndex("excel1.xls", 0))  # Sheet1
    # print(Excel_PO.getSheetNameByIndex("excel1.xls", 1))  # Sheet2
    # print(Excel_PO.getAllRows("excel1.xls"))  # 10  //10行数据
    # print(Excel_PO.getRowCol("excel1.xls", 0))  # [10, 4]  // Sheet1，10行4列
    print(Excel_PO.getRowCol("excel3.xlsx", "Sheet2"))  # [0, 0]  // Sheet2中没有数据
    print(Excel_PO.getRowValue("excel1.xls", 1, 0))  # [0.0, 0.0, 'yoyo', 43909.0]   //Sheet1,第二行数据，注意日期变成了43909.9
    print(Excel_PO.getRowValue("excel1.xls", 2, "Sheet1"))  # [1.0, 1.0, '', '2020.3.19']   //Sheet1,第三行数据，注意数字是浮点数，如1.0
    print(Excel_PO.getRowValue("excel1.xls", 0)[3])  # 3123213.0  //Sheet1,第一行第三列
    print(Excel_PO.getColValue("excel1.xls", 0, "Sheet1"))  # ['', 0.0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0]  // Sheet1的第一列
    print(Excel_PO.getColValue("excel1.xls", 0, 1))  # []

    # Excel_PO.writeXls("excel1.xls", "Sheet2", 2, 5, "令狐冲")  # 对Sheet2的第3行第6列写入令狐冲
    # Excel_PO.writeXls("excel1.xls", "Sheet2", "*", 5, "")  # 对Sheet2的第6列写入空白（清除） (保留第一行，一般第一行是标题)
    # Excel_PO.writeXls("excel1.xls", 1, 2, '*', 'aaa')  # 对Sheet2的第3行写入aaa (保留第一列，一般第一列是标题)

    # Excel_PO.writeXlsx("excel3.xlsx", "Sheet1", 5, 3, "测试一下")  # 对Sheet1的第5行第3列写入内容

    # Excel_PO.clearRowColForXlsx("excel3.xlsx", 1, "col", 1)  # Sheet2，清空Sheet2的第1列
    # Excel_PO.clearRowColForXlsx("excel3.xlsx", "Sheet2", "row", 2)  # 清空Sheet2的第2行。

    # Excel_PO.delRowColForXlsx("excel3.xlsx", "Sheet2", "col", 2, 1)  # 删除Sheet2的第2列
    # Excel_PO.delRowColForXlsx("excel3.xlsx", 1, "row", 3, 2)  # 删除Sheet2的第3行,第4行



