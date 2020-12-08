# coding: utf-8
# ********************************************************************************************************************
# Author     : John
# Date       : 2020-11-13
# Description: openpyxl 学习
# ********************************************************************************************************************

import openpyxl,sys

class Case: #这个类用来存储用例的
    __slots__ = [] #特殊的类属性，可以用来限制这个类创建的实例属性添加 可写可不写
    pass

class ReadExcel(object): #读取excel数据的类

    def __init__(self,file_name,sheet_name):
        """
        这个是用来初始化读取对象的
        :param file_name: 文件名 ---> str类型
        :param sheet_name: 表单名 ———> str类型
        """
        # 打开文件
        self.wb = openpyxl.load_workbook(file_name)
        # 选择表单
        self.sh = self.wb[sheet_name]

    def l_getRowData(self):

        ''' 获取每行数据 '''

        l_rowData = []  # 每行数据
        l_allData = []  # 所有行数据

        for cases in list(self.sh.rows):
            for i in range(len(cases)):
                l_rowData.append(cases[i].value)
            l_allData.append(l_rowData)
            l_rowData = []

        return (l_allData)

    def l_getRowDataByPartCol(self, l_varCol):

        '''
        获取指定列的行数据，保存到列表
        如：获取第1，3，4列的行数据
        '''

        l_rowData = []  # 每行的数据
        l_allData = []  # 所有的数据

        max_r = self.sh.max_row
        for row in range(1, max_r + 1):
            try:
                for column in l_varCol:
                    l_rowData.append(self.sh.cell(row, column).value)
                l_allData.append(l_rowData)
                l_rowData = []
            except:
                print("errorrrrrrrrrr, line " + str(sys._getframe(1).f_lineno) + ", in " + sys._getframe().f_code.co_name + "() ")
                print("建议：参数列表元素不能是0或负数")
                exit(0)

        return l_allData

    def l_getColData(self):

        ''' 获取每列数据 '''

        l_colData = []  # 每列数据
        l_allData = []  # 所有行数据

        for cases in list(self.sh.columns):
            for i in range(len(cases)):
                l_colData.append(cases[i].value)
            l_allData.append(l_colData)
            l_colData = []

        return (l_allData)

    def l_getColDataByPartCol(self, l_varCol):

        '''
        获取部分列的列数据，保存到列表
        如：获取第1，3，4列的列数据
        '''

        l_colData = []  # 每列的数据
        l_allData = []  # 所有的数据

        max_r = self.sh.max_row
        for col in l_varCol:
            try:
                for row in range(1, max_r + 1):
                    l_colData.append(self.sh.cell(row, col).value)
                l_allData.append(l_colData)
                l_colData = []
            except:
                print("errorrrrrrrrrr, line " + str(sys._getframe(1).f_lineno) + ", in " + sys._getframe().f_code.co_name + "() ")
                print("建议：参数列表元素不能是0或负数")
                exit(0)

        return l_allData



if __name__ == '__main__':

    r = ReadExcel('d:\\1.xlsx', 'Sheet')

    # l_data = r.l_getRowData()
    # print(l_data)
    #
    # l_data = r.l_getRowDataByPartCol([1,2,4])
    # print(l_data)

    l_data = r.l_getColData()
    print(l_data)
    for i in range(len(l_data[4])):
        print(l_data[4][i])

    l_data = r.l_getColDataByPartCol([1, 5])
    print(l_data)




# # wb = openpyxl.Workbook()
# # wb.create_sheet('test_case')
# # wb.save('d:\\cases.xlsx')
#
#
# wb = openpyxl.load_workbook('d:\\cases.xlsx')
# sh = wb['Sheet']
# ce = sh.cell(row=1, column=1)   # 读取第一行，第一列的数据
# # print(ce.value)
#
# # 将表格每行数据保存到列表里
# # print(list(sh.rows)[1:])     # 按行读取数据，去掉第一行的表头信息数据
# titles = []
# for title in (list(sh.rows)[0]):
#     titles.append(title.value)
# print(titles)
# list1 =[]
# list2 = []
# for cases in list(sh.rows)[1:]:
#     for i in range(len(cases)):
#         list1.append(cases[i].value)
#     list2.append(list1)
#     list1 = []
# print(list2)
#
# wb.close()