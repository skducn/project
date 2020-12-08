# -*- coding: utf-8 -*-
# *********************************************************************
# Author        : John
# Date          : 2019-9-26
# Description   : excel 对象层
# *********************************************************************

import openpyxl
import openpyxl.styles
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook("d:\\1.xlsx")
sheet_name = wb.sheetnames
sheet = wb[sheet_name[0]]  # 第一种方式，根据sheet表的坐标打开
# sheet = wb["python"]  # 第二种方式，直接指定sheet表的名称打开

print(wb.encoding)  # utf-8  # 获取文档的字符集编码
print(wb.properties) # 获取文档的元数据，如标题，创建者，创建日期等
# wb.active = 1  # 通过索引值设置当前活跃的worksheet


print(sheet.title)
print(sheet.max_row) # 表格的最大行数
print(sheet.min_row) # 表格的最小行数
print(sheet.max_column) # 表格的最大列数
print(sheet.min_column) # 表格的最小列数
print(sheet.rows) # 按行获取单元格(Cell对象) - 生成器
print(sheet.columns) # 按列获取单元格(Cell对象) - 生成器
print(sheet.values)# 按行获取表格的内容(数据)  - 生成器

# sheet.iter_rows() # 按行获取所有单元格，内置属性有(min_row,max_row,min_col,max_col)
# sheet.iter_cols() # 按列获取所有的单元格

# sheet["A"]  # 返回A列中所有的单元格cell对象
# sheet["1"]  # 返回第一行中所有的单元格cell对象
# sheet["A1"] # 返回该单元格对象



# 1，设置B1单元格字体为宋体，字号为17号，颜色为浅蓝色
sheet["B1"].font = openpyxl.styles.Font(name="宋体", size=33, color="00CCFF")

# 2，设置填充颜色为FFBB02,solid参数表示填充实色， fgColor=前景色，bgcolor=背景色
fille = PatternFill("solid", fgColor="FFBB02")
sheet.cell(1, 1).fill = fille
sheet.cell(4, 5).fill = fille

green_fill = PatternFill(fill_type="solid", fgColor="AACF91")
# 3，填充指定列的背景色，只对空白单元格有效（即单元格有内容或有颜色的不处理）
sheet.column_dimensions['A'].fill = green_fill

# 4，填充指定行的背景色，只对空白单元格有效（即单元格有内容或有颜色的不处理）
sheet.row_dimensions[4].fill = green_fill

# 5，将工作表标题选项卡的背景颜色设置为蓝色
sheet.sheet_properties.tabColor = "1072BA"

wb.save("d:\\1.xlsx")